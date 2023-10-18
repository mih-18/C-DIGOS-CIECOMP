import openpyxl

# Criar uma planilha
book = openpyxl.Workbook()

#Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet("Frutas")
#como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maçã', '12', 'R$6,50'])
frutas_page.append(['Uva', '25', 'R$12,90'])
frutas_page.append(['Abacaxi', '30', 'R$7,25'])
#Salvar a planilha
book.save('Planilha de Compras.xlsx')